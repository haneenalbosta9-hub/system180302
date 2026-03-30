"""
report_generator_excel.py
Professional Excel report generator for system1803 Microbiology Laboratory.

Replaces the Word-template approach to give full, reliable control over
font, bold, size, colour, borders, and layout — none of which were
preserved when python-docx replaced placeholder text in Word templates.

Usage (from app.py):
    from report_generator_excel import (
        generate_bioburden_report_excel,
        generate_sterility_report_excel,
        generate_endotoxin_report_excel,
        generate_environmental_report_excel,
    )

Every function returns an io.BytesIO object ready to pass to
st.download_button(..., data=buf, mime="application/vnd.openxmlformats-
officedocument.spreadsheetml.sheet").
"""

import io
import os

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

# ─── Colour palette ─────────────────────────────────────────────────────────
DARK_BLUE   = "1F4E79"
MID_BLUE    = "2E75B6"
LIGHT_BLUE  = "D6E4F0"
WHITE       = "FFFFFF"
LIGHT_GRAY  = "F2F2F2"
DARK_GRAY   = "595959"

LOGO_PATH = "logo.jpeg"   # relative to app working directory


# ─── Style helpers ───────────────────────────────────────────────────────────

def _side(style="thin", color=MID_BLUE):
    return Side(style=style, color=color)

def _border(style="thin", color=MID_BLUE):
    s = _side(style, color)
    return Border(left=s, right=s, top=s, bottom=s)

def _font(bold=False, size=10, color="000000", italic=False):
    return Font(name="Arial", bold=bold, size=size, color=color, italic=italic)

def _fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)

def _align(h="center", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _style_header_cell(cell, text):
    """Dark-blue header: white bold text, centred."""
    cell.value = text
    cell.font  = _font(bold=True, size=10, color=WHITE)
    cell.fill  = _fill(DARK_BLUE)
    cell.alignment = _align()
    cell.border = _border()

def _style_subheader_cell(cell, text):
    """Mid-blue subheader: white bold text, centred."""
    cell.value = text
    cell.font  = _font(bold=True, size=10, color=WHITE)
    cell.fill  = _fill(MID_BLUE)
    cell.alignment = _align()
    cell.border = _border()

def _style_label_cell(cell, text):
    """Light-blue label: dark-blue bold text, left-aligned."""
    cell.value = text
    cell.font  = _font(bold=True, size=10, color=DARK_BLUE)
    cell.fill  = _fill(LIGHT_BLUE)
    cell.alignment = _align(h="left")
    cell.border = _border()

def _style_value_cell(cell, text, bold=False):
    """White value cell: black text, left-aligned."""
    cell.value = str(text) if text is not None else ""
    cell.font  = _font(bold=bold, size=10)
    cell.alignment = _align(h="left")
    cell.border = _border()

def _style_data_cell(cell, text, alternating=False, center=True):
    """Standard data row cell."""
    cell.value = str(text) if text is not None else ""
    cell.font  = _font(size=10)
    cell.fill  = _fill(LIGHT_GRAY if alternating else WHITE)
    cell.alignment = _align(h="center" if center else "left")
    cell.border = _border()


# ─── Layout helpers ──────────────────────────────────────────────────────────

def _set_col_widths(ws, widths: dict):
    """widths: {'A': 25, 'B': 18, ...}"""
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _add_logo_header(ws, report_title: str, ncols: int) -> int:
    """
    Inserts logo (rows 1-3) + report title banner (row 4).
    Returns the next available row number (5).
    """
    last = get_column_letter(ncols)

    # ── Logo placeholder rows ─────────────────────────────────
    ws.row_dimensions[1].height = 55
    ws.row_dimensions[2].height = 8
    ws.row_dimensions[3].height = 8

    # Try inserting the actual logo image
    if os.path.exists(LOGO_PATH):
        try:
            img = XLImage(LOGO_PATH)
            img.width  = 140
            img.height = 60
            img.anchor = "A1"
            ws.add_image(img)
        except Exception:
            pass  # image load failed – continue without logo

    # Lab name (right of logo)
    ws.merge_cells(f"B1:{last}1")
    lab_cell = ws["B1"]
    lab_cell.value = "MICROBIOLOGY LABORATORY"
    lab_cell.font  = _font(bold=True, size=18, color=DARK_BLUE)
    lab_cell.alignment = _align(h="center")

    # ── Separator ─────────────────────────────────────────────
    ws.row_dimensions[3].height = 4
    ws.merge_cells(f"A3:{last}3")
    sep_cell = ws["A3"]
    sep_cell.fill = _fill(MID_BLUE)

    # ── Report title banner ───────────────────────────────────
    ws.row_dimensions[4].height = 28
    ws.merge_cells(f"A4:{last}4")
    title_cell = ws["A4"]
    title_cell.value     = report_title
    title_cell.font      = _font(bold=True, size=14, color=WHITE)
    title_cell.fill      = _fill(DARK_BLUE)
    title_cell.alignment = _align()
    title_cell.border    = _border(style="medium")

    return 5


def _add_info_table(ws, info_pairs: list, start_row: int, ncols: int) -> int:
    """
    Renders a 2-column label | value table.
    Label spans columns 1..mid, value spans mid+1..ncols.
    Returns the next available row after the table + a blank row.
    """
    mid     = ncols // 2
    mid_col = get_column_letter(mid)
    val_col = get_column_letter(mid + 1)
    last    = get_column_letter(ncols)

    for label, value in info_pairs:
        ws.row_dimensions[start_row].height = 20

        ws.merge_cells(f"A{start_row}:{mid_col}{start_row}")
        _style_label_cell(ws[f"A{start_row}"], label)

        ws.merge_cells(f"{val_col}{start_row}:{last}{start_row}")
        _style_value_cell(ws[f"{val_col}{start_row}"], value)

        start_row += 1

    # Blank separator row
    ws.row_dimensions[start_row].height = 8
    start_row += 1
    return start_row


def _add_section_banner(ws, text: str, row: int, ncols: int) -> int:
    last = get_column_letter(ncols)
    ws.row_dimensions[row].height = 22
    ws.merge_cells(f"A{row}:{last}{row}")
    _style_subheader_cell(ws[f"A{row}"], text)
    return row + 1


def _add_table_headers(ws, headers: list, row: int) -> int:
    for i, h in enumerate(headers, 1):
        _style_header_cell(ws.cell(row=row, column=i), h)
    ws.row_dimensions[row].height = 28
    return row + 1


def _add_footer(ws, row: int, ncols: int) -> int:
    last = get_column_letter(ncols)
    mid  = ncols // 2
    mid_col  = get_column_letter(mid)
    next_col = get_column_letter(mid + 1)

    # Blank gap
    ws.row_dimensions[row].height = 10
    row += 1

    # Disclaimer
    ws.merge_cells(f"A{row}:{last}{row}")
    disc = ws[f"A{row}"]
    disc.value = (
        "Results are valid for the tested samples only and shall not be "
        "reproduced except in full without written approval from the laboratory."
    )
    disc.font      = _font(italic=True, size=8, color=DARK_GRAY)
    disc.alignment = _align(h="center")
    ws.row_dimensions[row].height = 18
    row += 2

    # Prepared by / Approved by
    ws.merge_cells(f"A{row}:{mid_col}{row}")
    pb = ws[f"A{row}"]
    pb.value  = "Prepared by:"
    pb.font   = _font(bold=True, size=10)
    pb.border = Border(bottom=_side())
    pb.alignment = _align(h="left")

    ws.merge_cells(f"{next_col}{row}:{last}{row}")
    ab = ws[f"{next_col}{row}"]
    ab.value  = "Approved by:"
    ab.font   = _font(bold=True, size=10)
    ab.border = Border(bottom=_side())
    ab.alignment = _align(h="left")

    ws.row_dimensions[row].height = 40
    return row + 1


def _set_print_setup(ws):
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize   = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left   = 0.5
    ws.page_margins.right  = 0.5
    ws.page_margins.top    = 0.75
    ws.page_margins.bottom = 0.75
    ws.print_title_rows    = "1:4"   # repeat header on every printed page


def _wb_to_bytes(wb) -> io.BytesIO:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ════════════════════════════════════════════════════════════════════════════
#  1. BIOBURDEN REPORT
# ════════════════════════════════════════════════════════════════════════════

def generate_bioburden_report_excel(
    *,
    sample_id:        str,
    received_date:    str,
    test_date:        str,
    issuing_date:     str,
    customer_name:    str,
    sample_types:     list,   # list of str
    batch_list:       list,   # list of str
    reference_text:   str,
    tamc_text:        str,
    tymc_text:        str,
) -> io.BytesIO:
    """
    Returns an in-memory .xlsx file for the Bioburden report.

    Parameters
    ----------
    tamc_text : str  e.g. "No microbial growth was detected"  or  "120 CFU/ml"
    tymc_text : str  same format
    """
    NCOLS = 5
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bioburden Report"

    _set_col_widths(ws, {"A": 22, "B": 20, "C": 33, "D": 33, "E": 16})

    row = _add_logo_header(ws, "BIOBURDEN TEST REPORT", NCOLS)

    info = [
        ("Received Date:",        received_date),
        ("Test Performing Date:", test_date),
        ("Issuing Date:",         issuing_date),
        ("Customer Name:",        customer_name),
        ("Sample ID:",            sample_id),
        ("Sample Type:",          ", ".join(sample_types)),
        ("Sample Batch No.:",     " / ".join(batch_list)),
        ("Reference No.:",        reference_text),
    ]
    row = _add_info_table(ws, info, row, NCOLS)
    row = _add_section_banner(ws, "TEST RESULTS", row, NCOLS)

    headers = [
        "Sample ID",
        "Sample Batch No.",
        "Total Aerobic Microbial Count\n(TAMC) (CFU/ml)",
        "Total Combined Yeasts/Molds Count\n(TYMC) (CFU/ml)",
        "Status",
    ]
    row = _add_table_headers(ws, headers, row)

    for i, batch in enumerate(batch_list):
        alt = (i % 2 == 1)
        for col, val in enumerate([sample_id, batch, tamc_text, tymc_text, "Released"], 1):
            _style_data_cell(ws.cell(row=row, column=col), val, alt)
        ws.row_dimensions[row].height = 20
        row += 1

    _add_footer(ws, row + 1, NCOLS)
    _set_print_setup(ws)
    return _wb_to_bytes(wb)


# ════════════════════════════════════════════════════════════════════════════
#  2. STERILITY REPORT
# ════════════════════════════════════════════════════════════════════════════

def generate_sterility_report_excel(
    *,
    sample_id:      str,
    received_date:  str,
    test_date:      str,
    issuing_date:   str,
    customer_name:  str,
    sample_types:   list,
    batch_list:     list,
    reference_text: str,
    result_text:    str,   # full result paragraph(s)
) -> io.BytesIO:
    NCOLS = 4
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sterility Report"

    _set_col_widths(ws, {"A": 25, "B": 25, "C": 25, "D": 25})

    row = _add_logo_header(ws, "STERILITY TEST REPORT", NCOLS)

    info = [
        ("Received Date:",        received_date),
        ("Test Performing Date:", test_date),
        ("Issuing Date:",         issuing_date),
        ("Customer Name:",        customer_name),
        ("Sample ID:",            sample_id),
        ("Sample Type:",          ", ".join(sample_types)),
        ("Sample Batch No.:",     " / ".join(batch_list)),
        ("Reference No.:",        reference_text),
    ]
    row = _add_info_table(ws, info, row, NCOLS)
    row = _add_section_banner(ws, "TEST RESULTS", row, NCOLS)

    # Multi-line result block
    result_lines = result_text.count("\n") + 1
    result_height = max(4, result_lines + 1)
    ws.merge_cells(f"A{row}:D{row + result_height - 1}")
    rc = ws[f"A{row}"]
    rc.value     = result_text
    rc.font      = _font(size=10)
    rc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    rc.border    = _border()
    for r in range(row, row + result_height):
        ws.row_dimensions[r].height = 20
    row += result_height + 1

    _add_footer(ws, row + 1, NCOLS)
    _set_print_setup(ws)
    return _wb_to_bytes(wb)


# ════════════════════════════════════════════════════════════════════════════
#  3. ENDOTOXIN REPORT
# ════════════════════════════════════════════════════════════════════════════

def generate_endotoxin_report_excel(
    *,
    sample_id:        str,
    received_date:    str,
    test_date:        str,
    issuing_date:     str,
    customer_name:    str,
    sample_rows_data: list,   # list of dicts: {sample_type, batch_no, ref_no}
    endotoxin_result: str,    # e.g. "<0.01 EU/ml"
) -> io.BytesIO:
    NCOLS = 5
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Endotoxin Report"

    _set_col_widths(ws, {"A": 28, "B": 18, "C": 22, "D": 22, "E": 22})

    row = _add_logo_header(ws, "ENDOTOXIN TEST REPORT", NCOLS)

    info = [
        ("Received Date:",        received_date),
        ("Test Performing Date:", test_date),
        ("Issuing Date:",         issuing_date),
        ("Customer Name:",        customer_name),
        ("Sample ID:",            sample_id),
    ]
    row = _add_info_table(ws, info, row, NCOLS)
    row = _add_section_banner(ws, "TEST RESULTS", row, NCOLS)

    headers = [
        "Product Name / Sample Type",
        "Sample ID",
        "Batch Number",
        "Reference No.",
        "Endotoxin Level (EU/ml)",
    ]
    row = _add_table_headers(ws, headers, row)

    for i, r in enumerate(sample_rows_data):
        alt = (i % 2 == 1)
        vals = [
            r.get("sample_type", ""),
            sample_id,
            r.get("batch_no", ""),
            r.get("ref_no", ""),
            endotoxin_result,
        ]
        for col, val in enumerate(vals, 1):
            _style_data_cell(ws.cell(row=row, column=col), val, alt)
        ws.row_dimensions[row].height = 20
        row += 1

    _add_footer(ws, row + 1, NCOLS)
    _set_print_setup(ws)
    return _wb_to_bytes(wb)


# ════════════════════════════════════════════════════════════════════════════
#  4. ENVIRONMENTAL REPORT
# ════════════════════════════════════════════════════════════════════════════

def generate_environmental_report_excel(
    *,
    received_date:     str,
    test_date:         str,
    issuing_date:      str,
    customer_name_ar:  str,
    customer_name_en:  str,
    sample_type:       str,
    results_data:      list,        # list of dicts: {NO, Customer ID, Fungus Count (CFU), Total Bacterial Count (CFU)}
    product_names_map: dict,        # {sample_id: product_name}
) -> io.BytesIO:
    NCOLS = 6
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Environmental Report"

    _set_col_widths(ws, {"A": 7, "B": 22, "C": 32, "D": 14, "E": 20, "F": 22})

    row = _add_logo_header(ws, "ENVIRONMENTAL MONITORING REPORT", NCOLS)

    info = [
        ("Received Date:",        received_date),
        ("Test Performing Date:", test_date),
        ("Issuing Date:",         issuing_date),
        ("Customer (Arabic):",    customer_name_ar),
        ("Customer (English):",   customer_name_en),
        ("Sample Type:",          sample_type),
    ]
    row = _add_info_table(ws, info, row, NCOLS)

    # ── Sample / Product info table ───────────────────────────
    row = _add_section_banner(ws, "SAMPLE INFORMATION", row, NCOLS)

    # Header: No. | Sample ID | Product Name/Location (spans C-F)
    for col, txt in enumerate(["No.", "Sample ID"], 1):
        _style_header_cell(ws.cell(row=row, column=col), txt)
    ws.merge_cells(f"C{row}:F{row}")
    _style_header_cell(ws[f"C{row}"], "Product Name / Location")
    ws.row_dimensions[row].height = 22
    row += 1

    for i, res in enumerate(results_data):
        sid = res["Customer ID"]
        alt = (i % 2 == 1)
        _style_data_cell(ws.cell(row=row, column=1), res["NO"],  alt)
        _style_data_cell(ws.cell(row=row, column=2), sid,        alt)
        ws.merge_cells(f"C{row}:F{row}")
        _style_data_cell(ws["C" + str(row)], product_names_map.get(sid, ""), alt, center=False)
        ws.row_dimensions[row].height = 18
        row += 1

    row += 1  # gap

    # ── Results table ─────────────────────────────────────────
    row = _add_section_banner(ws, "TEST RESULTS", row, NCOLS)

    # Two-line spanning headers
    header_map = [
        (1, 1, "No."),
        (2, 2, "Sample ID"),
        (3, 4, "Product Name / Location"),   # merged C:D
        (5, 5, "Fungus Count\n(CFU/plate)"),
        (6, 6, "Total Bacterial Count\n(CFU/plate)"),
    ]
    for start_c, end_c, txt in header_map:
        if start_c == end_c:
            _style_header_cell(ws.cell(row=row, column=start_c), txt)
        else:
            ws.merge_cells(
                start_row=row, end_row=row,
                start_column=start_c, end_column=end_c
            )
            _style_header_cell(ws.cell(row=row, column=start_c), txt)
    ws.row_dimensions[row].height = 35
    row += 1

    for i, res in enumerate(results_data):
        sid = res["Customer ID"]
        alt = (i % 2 == 1)
        _style_data_cell(ws.cell(row=row, column=1), res["NO"],  alt)
        _style_data_cell(ws.cell(row=row, column=2), sid,        alt)

        ws.merge_cells(f"C{row}:D{row}")
        _style_data_cell(ws["C" + str(row)],
                         product_names_map.get(sid, ""), alt, center=False)

        _style_data_cell(ws.cell(row=row, column=5),
                         res.get("Fungus Count (CFU)", ""),          alt)
        _style_data_cell(ws.cell(row=row, column=6),
                         res.get("Total Bacterial Count (CFU)", ""), alt)
        ws.row_dimensions[row].height = 18
        row += 1

    _add_footer(ws, row + 1, NCOLS)
    _set_print_setup(ws)
    return _wb_to_bytes(wb)
